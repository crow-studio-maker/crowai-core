"""Safe local document extraction for Agent V1.0 attachments."""

from __future__ import annotations

import base64
import bz2
import csv
import gzip
import html
import io
import json
import lzma
import mimetypes
import re
import tarfile
import zipfile
from html.parser import HTMLParser
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree


class DocumentExtractionError(RuntimeError):
    """Raised when a supported document cannot be inspected safely."""


_TEXT_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".rst", ".log", ".py", ".pyi", ".pyw", ".js", ".jsx", ".mjs", ".cjs",
    ".ts", ".tsx", ".css", ".scss", ".less", ".html", ".htm", ".json", ".jsonl", ".csv", ".tsv",
    ".xml", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".conf", ".sql", ".java", ".c", ".h", ".cpp",
    ".hpp", ".cc", ".cxx", ".cs", ".go", ".rs", ".rb", ".php", ".swift", ".kt", ".kts", ".scala",
    ".sh", ".bash", ".zsh", ".fish", ".bat", ".cmd", ".ps1", ".tex", ".vue", ".svelte", ".dart",
    ".gradle", ".properties", ".lock", ".graphql", ".proto", ".r", ".lua", ".pl", ".asm", ".s", ".cmake",
    ".make", ".mk", ".env",
}

_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"}
_OFFICE_EXTENSIONS = {".docx", ".xlsx", ".xlsm", ".pptx", ".odt", ".ods", ".odp"}
_ARCHIVE_EXTENSIONS = {".zip", ".jar", ".whl", ".epub", ".apk", ".ipa", ".vsix"}
_MAX_GENERIC_BYTES = 4 * 1024 * 1024


class _VisibleHtmlParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self._skip = 0
        self.parts: list[str] = []

    def handle_starttag(
        self,
        tag: str,
        attrs: list[tuple[str, str | None]],
    ) -> None:
        if tag.casefold() in {"script", "style", "noscript", "svg"}:
            self._skip += 1

    def handle_endtag(self, tag: str) -> None:
        if (
            tag.casefold() in {"script", "style", "noscript", "svg"}
            and self._skip > 0
        ):
            self._skip -= 1

    def handle_data(self, data: str) -> None:
        if self._skip:
            return

        value = " ".join(data.split())

        if value:
            self.parts.append(value)


def _decode(raw: bytes) -> str:
    for encoding in (
        "utf-8",
        "utf-8-sig",
        "cp1254",
        "cp1252",
        "latin-1",
    ):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue

    return raw.decode("utf-8", errors="replace")


def _normalize_text(value: str, maximum_chars: int) -> str:
    text = str(value or "").replace("\x00", " ")
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{4,}", "\n\n\n", text)
    return text.strip()[:maximum_chars]


def _safe_archive_name(value: str) -> str | None:
    path = PurePosixPath(str(value).replace("\\", "/"))

    if path.is_absolute():
        return None

    if any(part in {"", ".", ".."} for part in path.parts):
        return None

    return path.as_posix()



def _validate_zip_archive(
    path: Path,
    *,
    maximum_members: int,
    maximum_uncompressed_bytes: int,
) -> None:
    """Reject encrypted, path-unsafe and decompression-bomb-like ZIP members."""

    with zipfile.ZipFile(path) as archive:
        members = archive.infolist()
        if len(members) > maximum_members:
            raise DocumentExtractionError("Archive contains too many members for safe inspection.")
        total = 0
        for member in members:
            safe_name = _safe_archive_name(member.filename)
            unix_mode = (member.external_attr >> 16) & 0o170000
            if not safe_name or member.flag_bits & 0x1 or unix_mode == 0o120000:
                raise DocumentExtractionError("Archive contains an unsafe or encrypted member.")
            if member.file_size > maximum_uncompressed_bytes:
                raise DocumentExtractionError("Archive contains an oversized member.")
            if member.file_size and member.compress_size == 0:
                raise DocumentExtractionError("Archive contains an invalid compressed member.")
            if member.compress_size and member.file_size / member.compress_size > 250:
                raise DocumentExtractionError("Archive compression ratio exceeds the safety limit.")
            total += member.file_size
            if total > maximum_uncompressed_bytes:
                raise DocumentExtractionError("Archive expands beyond the safety limit.")


def _zip_container_kind(path: Path) -> tuple[str, str]:
    """Recognize ZIP-based office documents independently of filename suffix."""

    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        if "word/document.xml" in names:
            return "docx", "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        if "ppt/presentation.xml" in names:
            return "pptx", "application/vnd.openxmlformats-officedocument.presentationml.presentation"
        if "xl/workbook.xml" in names:
            return "xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if "content.xml" in names and "META-INF/manifest.xml" in names:
            media = "application/vnd.oasis.opendocument.text"
            try:
                value = archive.read("mimetype")[:160].decode("ascii", errors="ignore").strip()
                if value.startswith("application/vnd.oasis.opendocument."):
                    media = value
            except KeyError:
                pass
            return "opendocument", media
    return "zip", "application/zip"


def _magic_media_type(path: Path) -> str:
    try:
        with path.open("rb") as handle:
            sample = handle.read(16)
    except OSError:
        return ""
    if sample.startswith(b"%PDF-"):
        return "application/pdf"
    if sample.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if sample.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    if sample.startswith((b"GIF87a", b"GIF89a")):
        return "image/gif"
    if sample.startswith(b"BM"):
        return "image/bmp"
    if sample.startswith(b"RIFF") and sample[8:12] == b"WEBP":
        return "image/webp"
    if sample.startswith(b"PK\x03\x04"):
        return "application/zip"
    if sample.startswith(b"\x1f\x8b"):
        return "application/gzip"
    if sample.startswith(b"BZh"):
        return "application/x-bzip2"
    if sample.startswith(b"\xfd7zXZ\x00"):
        return "application/x-xz"
    if sample.startswith(b"7z\xbc\xaf\x27\x1c"):
        return "application/x-7z-compressed"
    if sample.startswith(b"Rar!\x1a\x07"):
        return "application/vnd.rar"
    if sample.startswith(b"MZ"):
        return "application/vnd.microsoft.portable-executable"
    if sample.startswith(b"\x7fELF"):
        return "application/x-elf"
    return ""

def _xml_text(raw: bytes) -> str:
    try:
        root = ElementTree.fromstring(raw)
    except ElementTree.ParseError:
        return ""

    values: list[str] = []

    for element in root.iter():
        local_name = element.tag.rsplit("}", 1)[-1]

        if local_name in {"t", "v"} and element.text:
            values.append(element.text)

        if local_name in {"p", "tr"}:
            values.append("\n")

    return " ".join(values).replace(" \n ", "\n")


def _extract_docx(path: Path, maximum_chars: int) -> str:
    with zipfile.ZipFile(path) as archive:
        names = [
            name
            for name in archive.namelist()
            if name.startswith("word/")
            and name.endswith(".xml")
            and (
                name == "word/document.xml"
                or "/header" in name
                or "/footer" in name
                or "footnotes" in name
                or "endnotes" in name
            )
        ]

        parts = [
            _xml_text(archive.read(name))
            for name in names
        ]

    return _normalize_text("\n\n".join(parts), maximum_chars)


def _extract_pptx(path: Path, maximum_chars: int) -> str:
    with zipfile.ZipFile(path) as archive:
        slide_names = sorted(
            name
            for name in archive.namelist()
            if name.startswith("ppt/slides/slide")
            and name.endswith(".xml")
        )
        parts: list[str] = []

        for index, name in enumerate(slide_names, start=1):
            parts.append(
                f"[Slide {index}]\n{_xml_text(archive.read(name))}"
            )

    return _normalize_text("\n\n".join(parts), maximum_chars)


def _extract_xlsx(path: Path, maximum_chars: int) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        load_workbook = None

    if load_workbook is not None:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=True,
        )
        parts: list[str] = []
        characters = 0

        try:
            for worksheet in workbook.worksheets:
                parts.append(f"[Sheet: {worksheet.title}]")

                for row_index, row in enumerate(
                    worksheet.iter_rows(values_only=True),
                    start=1,
                ):
                    if row_index > 5000:
                        parts.append("[Sheet truncated after 5000 rows]")
                        break

                    values = [
                        "" if value is None else str(value)
                        for value in row[:100]
                    ]
                    line = "\t".join(values).rstrip()

                    if line:
                        parts.append(line)
                        characters += len(line)

                    if characters >= maximum_chars:
                        return _normalize_text(
                            "\n".join(parts),
                            maximum_chars,
                        )
        finally:
            workbook.close()

        return _normalize_text("\n".join(parts), maximum_chars)

    with zipfile.ZipFile(path) as archive:
        parts = [
            _xml_text(archive.read(name))
            for name in archive.namelist()
            if name.startswith("xl/worksheets/")
            and name.endswith(".xml")
        ]

    return _normalize_text("\n".join(parts), maximum_chars)


def _embedded_zip_images(
    path: Path,
    *,
    prefixes: tuple[str, ...],
    maximum_images: int = 4,
    maximum_image_bytes: int = 8 * 1024 * 1024,
) -> list[dict[str, str]]:
    """Collect a few package-owned embedded images as bounded data URLs."""

    output: list[dict[str, str]] = []
    with zipfile.ZipFile(path) as archive:
        for member in archive.infolist():
            if len(output) >= maximum_images:
                break
            safe_name = _safe_archive_name(member.filename)
            if not safe_name or member.is_dir():
                continue
            if not any(safe_name.startswith(prefix) for prefix in prefixes):
                continue
            suffix = Path(safe_name).suffix.casefold()
            media_type = {
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
                ".png": "image/png",
                ".webp": "image/webp",
                ".gif": "image/gif",
                ".bmp": "image/bmp",
            }.get(suffix)
            if not media_type or member.file_size <= 0 or member.file_size > maximum_image_bytes:
                continue
            raw = archive.read(member)
            if len(raw) > maximum_image_bytes:
                continue
            output.append({
                "name": f"{path.name} — {safe_name}",
                "media_type": media_type,
                "data_url": f"data:{media_type};base64,{base64.b64encode(raw).decode('ascii')}",
            })
    return output


def _render_pdf_pages_for_vision(
    path: Path,
    *,
    maximum_pages: int = 4,
) -> list[dict[str, str]]:
    """Render a few PDF pages to in-memory PNG data URLs for the VL model."""

    try:
        import fitz
    except ImportError:
        return []

    output: list[dict[str, str]] = []
    try:
        document = fitz.open(str(path))
    except Exception:
        return []

    try:
        for index in range(min(len(document), maximum_pages)):
            page = document[index]
            pixmap = page.get_pixmap(matrix=fitz.Matrix(1.35, 1.35), alpha=False)
            png = pixmap.tobytes("png")
            if len(png) > 8 * 1024 * 1024:
                continue
            encoded = base64.b64encode(png).decode("ascii")
            output.append({
                "name": f"{path.name} — page {index + 1}",
                "media_type": "image/png",
                "data_url": f"data:image/png;base64,{encoded}",
            })
    except Exception:
        return output
    finally:
        document.close()

    return output


def _extract_pdf(
    path: Path,
    maximum_chars: int,
    maximum_pages: int,
) -> tuple[str, int, list[dict[str, str]]]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise DocumentExtractionError(
            "PDF support requires pypdf. Run: "
            "pip install -r requirements-agent.txt"
        ) from exc

    try:
        reader = PdfReader(str(path), strict=False)
    except Exception as exc:
        raise DocumentExtractionError(
            f"PDF could not be opened: {exc}"
        ) from exc

    parts: list[str] = []
    used_pages = min(len(reader.pages), maximum_pages)
    characters = 0

    for index in range(used_pages):
        try:
            page_text = reader.pages[index].extract_text() or ""
        except Exception as exc:
            page_text = f"[Page extraction failed: {exc}]"

        block = f"[Page {index + 1}]\n{page_text.strip()}"
        parts.append(block)
        characters += len(block)

        if characters >= maximum_chars:
            break

    normalized = _normalize_text(
        "\n\n".join(parts),
        maximum_chars,
    )
    meaningful = re.sub(
        r"\[Page \d+\]|\s+",
        "",
        normalized,
    )
    derived_images = (
        _render_pdf_pages_for_vision(path)
        if len(meaningful) < 120
        else []
    )

    return (
        normalized,
        len(reader.pages),
        derived_images,
    )


def _extract_zip(
    path: Path,
    *,
    maximum_chars: int,
    maximum_members: int,
    maximum_uncompressed_bytes: int,
) -> tuple[str, list[str]]:
    text_parts: list[str] = []
    names: list[str] = []
    total_uncompressed = 0

    with zipfile.ZipFile(path) as archive:
        members = archive.infolist()[:maximum_members]

        for member in members:
            safe_name = _safe_archive_name(member.filename)

            if not safe_name or member.is_dir():
                continue

            total_uncompressed += member.file_size

            if total_uncompressed > maximum_uncompressed_bytes:
                text_parts.append(
                    "[Archive text extraction stopped at safety limit.]"
                )
                break

            names.append(safe_name)
            suffix = Path(safe_name).suffix.casefold()

            if suffix not in _TEXT_EXTENSIONS:
                continue

            raw = archive.read(member)
            value = _decode(raw)

            if suffix in {".html", ".htm"}:
                parser = _VisibleHtmlParser()
                parser.feed(value)
                value = "\n".join(parser.parts)

            text_parts.append(
                f"[Archive member: {safe_name}]\n{value}"
            )

            if sum(len(item) for item in text_parts) >= maximum_chars:
                break

    return (
        _normalize_text("\n\n".join(text_parts), maximum_chars),
        names,
    )



def _extract_opendocument(path: Path, maximum_chars: int) -> str:
    with zipfile.ZipFile(path) as archive:
        content = archive.read("content.xml")
    try:
        root = ElementTree.fromstring(content)
    except ElementTree.ParseError as exc:
        raise DocumentExtractionError("OpenDocument content.xml is invalid.") from exc
    parts = [" ".join((element.text or "").split()) for element in root.iter() if element.text]
    return _normalize_text("\n".join(part for part in parts if part), maximum_chars)


def _extract_tar(
    path: Path,
    *,
    maximum_chars: int,
    maximum_members: int,
    maximum_uncompressed_bytes: int,
) -> tuple[str, list[str]]:
    parts: list[str] = []
    names: list[str] = []
    total = 0
    with tarfile.open(path, mode="r:*") as archive:
        members = archive.getmembers()
        if len(members) > maximum_members:
            members = members[:maximum_members]
        for member in members:
            safe_name = _safe_archive_name(member.name)
            if not safe_name or not member.isfile() or member.issym() or member.islnk():
                continue
            total += member.size
            if total > maximum_uncompressed_bytes:
                parts.append("[Archive inspection stopped at safety limit.]")
                break
            names.append(safe_name)
            suffix = Path(safe_name).suffix.casefold()
            if suffix not in _TEXT_EXTENSIONS or member.size > 1_000_000:
                continue
            handle = archive.extractfile(member)
            if handle is None:
                continue
            parts.append(f"[Archive member: {safe_name}]\n{_decode(handle.read(1_000_001))}")
            if sum(len(item) for item in parts) >= maximum_chars:
                break
    return _normalize_text("\n\n".join(parts), maximum_chars), names


def _binary_strings(raw: bytes, maximum_chars: int) -> str:
    values: list[str] = []
    seen: set[str] = set()
    candidates: list[str] = [
        match.group().decode("ascii", errors="ignore")
        for match in re.finditer(rb"[\x20-\x7e]{5,}", raw)
    ]
    for match in re.finditer(rb"(?:[\x20-\x7e]\x00){5,}", raw):
        try:
            candidates.append(match.group().decode("utf-16-le"))
        except UnicodeDecodeError:
            continue
    characters = 0
    for candidate in candidates:
        value = " ".join(candidate.split())[:500]
        key = value.casefold()
        if value and key not in seen:
            seen.add(key)
            values.append(value)
            characters += len(value)
        if len(values) >= 600 or characters >= maximum_chars:
            break
    return _normalize_text("\n".join(values), maximum_chars)

def _decompressed_text(path: Path, suffix: str, maximum_chars: int) -> str:
    opener = gzip.open if suffix == ".gz" else bz2.open if suffix == ".bz2" else lzma.open
    with opener(path, "rb") as handle:
        raw = handle.read(min(_MAX_GENERIC_BYTES, maximum_chars * 4))
    return _normalize_text(_decode(raw), maximum_chars)

def inspect_document(
    *,
    path: Path,
    original_name: str,
    media_type: str = "",
    maximum_chars: int = 240000,
    maximum_pages: int = 250,
    maximum_archive_members: int = 100,
    maximum_archive_bytes: int = 25000000,
) -> dict[str, Any]:
    """Inspect one user-supplied file without executing its contents."""

    source = Path(path).resolve()

    if not source.is_file():
        return {
            "status": "error",
            "text": "",
            "summary": "The uploaded file could not be found.",
        }

    name = str(original_name or source.name)
    suffix = Path(name).suffix.casefold()
    magic_media = _magic_media_type(source)
    supplied_media = str(media_type or "").strip()
    resolved_media = (
        magic_media
        or supplied_media
        or mimetypes.guess_type(name)[0]
        or "application/octet-stream"
    )
    result: dict[str, Any] = {
        "name": name,
        "path": str(source),
        "media_type": resolved_media,
        "detected_format": "unknown",
        "size_bytes": source.stat().st_size,
        "text": "",
        "status": "stored",
        "summary": "File stored.",
    }

    if (
        suffix in _IMAGE_EXTENSIONS
        or resolved_media.startswith("image/")
    ):
        result.update(
            {
                "detected_format": resolved_media.split("/", 1)[-1] if resolved_media.startswith("image/") else suffix.lstrip("."),
                "status": "image_ready",
                "summary": (
                    "Image is ready for local multimodal analysis."
                ),
                "image_path": str(source),
            }
        )
        return result

    try:
        is_zip = zipfile.is_zipfile(source)
        zip_kind = ""
        if is_zip:
            _validate_zip_archive(
                source,
                maximum_members=maximum_archive_members,
                maximum_uncompressed_bytes=maximum_archive_bytes,
            )
            zip_kind, zip_media = _zip_container_kind(source)
            if zip_kind != "zip":
                resolved_media = zip_media
                result["media_type"] = zip_media
            result["detected_format"] = zip_kind
        elif resolved_media == "application/pdf":
            result["detected_format"] = "pdf"
        elif resolved_media.startswith("image/"):
            result["detected_format"] = resolved_media.split("/", 1)[-1]

        if suffix == ".pdf" or resolved_media == "application/pdf":
            (
                text,
                total_pages,
                derived_images,
            ) = _extract_pdf(
                source,
                maximum_chars,
                maximum_pages,
            )
            result["page_count"] = total_pages
            result["derived_images"] = derived_images

        elif suffix == ".docx" or zip_kind == "docx":
            text = _extract_docx(source, maximum_chars)
            result["derived_images"] = _embedded_zip_images(
                source, prefixes=("word/media/",),
            )

        elif suffix == ".pptx" or zip_kind == "pptx":
            text = _extract_pptx(source, maximum_chars)
            result["derived_images"] = _embedded_zip_images(
                source, prefixes=("ppt/media/",),
            )

        elif suffix in {".xlsx", ".xlsm"} or zip_kind == "xlsx":
            text = _extract_xlsx(source, maximum_chars)
            result["derived_images"] = _embedded_zip_images(
                source, prefixes=("xl/media/",),
            )

        elif suffix in {".odt", ".ods", ".odp"} or zip_kind == "opendocument":
            text = _extract_opendocument(source, maximum_chars)
            result["derived_images"] = _embedded_zip_images(
                source, prefixes=("Pictures/",),
            )

        elif suffix in _ARCHIVE_EXTENSIONS or is_zip:
            text, members = _extract_zip(
                source,
                maximum_chars=maximum_chars,
                maximum_members=maximum_archive_members,
                maximum_uncompressed_bytes=maximum_archive_bytes,
            )
            result["archive_members"] = members

        elif tarfile.is_tarfile(source):
            text, members = _extract_tar(
                source,
                maximum_chars=maximum_chars,
                maximum_members=maximum_archive_members,
                maximum_uncompressed_bytes=maximum_archive_bytes,
            )
            result["archive_members"] = members

        elif suffix in {".gz", ".bz2", ".xz", ".lzma"} or resolved_media in {
            "application/gzip", "application/x-bzip2", "application/x-xz",
        }:
            compression_suffix = suffix if suffix in {".gz", ".bz2", ".xz", ".lzma"} else {
                "application/gzip": ".gz",
                "application/x-bzip2": ".bz2",
                "application/x-xz": ".xz",
            }[resolved_media]
            text = _decompressed_text(source, compression_suffix, maximum_chars)
            result["detected_format"] = compression_suffix.lstrip(".")

        elif suffix in _TEXT_EXTENSIONS or resolved_media.startswith("text/"):
            raw = source.read_bytes()[:_MAX_GENERIC_BYTES]
            text = _decode(raw)

            if suffix in {".html", ".htm"}:
                parser = _VisibleHtmlParser()
                parser.feed(text)
                text = "\n".join(parser.parts)

            elif suffix == ".json":
                try:
                    value = json.loads(text)
                    text = json.dumps(
                        value,
                        ensure_ascii=False,
                        indent=2,
                    )
                except json.JSONDecodeError:
                    pass

            elif suffix in {".csv", ".tsv"}:
                delimiter = "\t" if suffix == ".tsv" else ","
                rows = csv.reader(
                    io.StringIO(text),
                    delimiter=delimiter,
                )
                text = "\n".join(
                    "\t".join(row[:100])
                    for _, row in zip(range(5000), rows)
                )

            text = _normalize_text(text, maximum_chars)

        else:
            raw = source.read_bytes()[:_MAX_GENERIC_BYTES]
            printable = sum(byte in b"\n\r\t" or 32 <= byte < 127 for byte in raw)
            if raw and printable / len(raw) > 0.80:
                text = _normalize_text(_decode(raw), maximum_chars)
            else:
                strings = _binary_strings(raw, min(maximum_chars, 36000))
                text = ("Printable strings extracted from binary data:\n" + strings) if strings else ""
                result["binary_inspection"] = True

    except (
        OSError,
        zipfile.BadZipFile,
        tarfile.TarError,
        DocumentExtractionError,
        ValueError,
    ) as exc:
        result.update(
            {
                "status": "error",
                "summary": str(exc),
                "error": str(exc),
            }
        )
        return result

    derived_count = len(
        result.get("derived_images", [])
        if isinstance(result.get("derived_images"), list)
        else []
    )
    meaningful_text = bool(
        re.sub(r"\[Page \d+\]|\s+", "", text)
    )

    result.update(
        {
            "text": text,
            "status": (
                "inspected"
                if meaningful_text
                else (
                    "image_ready"
                    if derived_count
                    else "stored"
                )
            ),
            "summary": (
                f"Document inspected: {len(text)} characters extracted."
                if meaningful_text
                else (
                    f"No usable text layer was found; {derived_count} "
                    "visual item(s) were prepared for multimodal analysis."
                    if derived_count
                    else (
                        "File stored safely; no extractable text or printable strings were found."
                    )
                )
            ),
            "truncated": len(text) >= maximum_chars,
        }
    )
    return result
